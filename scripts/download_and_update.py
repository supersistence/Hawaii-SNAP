#!/usr/bin/env python3
"""
Hawaii SNAP Data Download and Update Script
============================================

This script automates the download and processing of updated SNAP data for Hawaii.

Usage:
    python download_and_update.py --all
    python download_and_update.py --monthly
    python download_and_update.py --retailers
    python download_and_update.py --county

Requirements:
    pip install pandas openpyxl requests

Data Sources:
    1. Statewide Monthly Data: USDA FNS SNAP Data Tables
    2. Historical Retailer Data: USDA FNS Retailer Historical Data
    3. County Bi-Annual Data: USDA FNS SNAP Data Tables
    4. State Application Data: Hawaii DHS
"""

import argparse
import os
import sys
import zipfile
from pathlib import Path
from datetime import datetime
import requests
import pandas as pd

# Configuration
DATA_DIR = Path(__file__).parent.parent / "Data"
BACKUP_DIR = DATA_DIR / "backups"
SOURCES_PATH = DATA_DIR / "SOURCES.json"  # provenance manifest (committed)

# Landing pages that link the actual data files. We scrape these for the
# current resource-file URL so version bumps (…-6.zip → …-7.zip) don't break us.
URLS = {
    "monthly": "https://www.fns.usda.gov/pd/supplemental-nutrition-assistance-program-snap",
    "retailers": "https://www.fns.usda.gov/snap/retailer/historical-data",
    "county": "https://www.fns.usda.gov/pd/supplemental-nutrition-assistance-program-snap",
    "dhs": "https://humanservices.hawaii.gov/bessd/snap/"
}

# Substring that identifies each dataset's file among the page's resource links.
SOURCE_FILE_PATTERNS = {
    "monthly": "snap-zip-fy69tocurrent",      # ZIP of per-FY Excel files (FY69..current)
    "retailers": "retailer-locator-data",     # Historical SNAP Retailer Locator ZIP
}

# Hawaii coordinate bounds, used to flag bad geolocation in retailer data.
HI_BOUNDS = {"lat": (18.9, 22.2), "lon": (-160.2, -154.8)}

# Akamai fronts fns.usda.gov and rejects requests without a browser UA.
USER_AGENT = "Mozilla/5.0 (compatible; Hawaii-SNAP-data-updater/1.0)"

# Expected file names (current naming)
FILES = {
    "monthly": "Statewide Monthly SNAP FY 89-25.csv",
    "retailers": "hawaii_snap_retailers_2004-2024_valid_coords.csv",
    "county": "County Bi-Annual SNAP 89-25.csv",
    "applications": "County Weekly Applications 4:2020-3:2022.csv"
}


def backup_existing_data(file_path):
    """Create a timestamped backup of existing data file."""
    if not file_path.exists():
        print(f"No existing file to backup: {file_path}")
        return

    BACKUP_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{file_path.stem}_{timestamp}{file_path.suffix}"
    backup_path = BACKUP_DIR / backup_name

    import shutil
    shutil.copy2(file_path, backup_path)
    print(f"✓ Backed up to: {backup_path}")


def read_asof_date(fy_file_path):
    """Read USDA's 'data as of' date from an FY file's title row (or None)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fy_file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    return cell.strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def write_provenance(key, entry):
    """Merge one dataset's provenance into the committed Data/SOURCES.json."""
    import json
    try:
        manifest = json.loads(SOURCES_PATH.read_text()) if SOURCES_PATH.exists() else {}
    except Exception:
        manifest = {}
    # Drop None-valued fields so a routine pull can refresh volatile fields
    # (downloadedAt, coverageEnd, ...) without clobbering curated descriptive
    # fields (note, fiscalYear) set by the backfill.
    entry = {k: v for k, v in entry.items() if v is not None}
    manifest[key] = {**manifest.get(key, {}), **entry}
    SOURCES_PATH.write_text(json.dumps(manifest, indent=2) + "\n")
    print(f"✓ Recorded provenance in {SOURCES_PATH.name} "
          f"(revision {entry.get('revision', '?')}, "
          f"as of {entry.get('dataAsOf', '?')})")


def discover_file_url(page_url, name_contains):
    """Scrape a USDA landing page for the current resource-file link.

    Returns an absolute URL whose path contains `name_contains`, or None.
    This keeps us resilient to USDA's file-revision suffixes (…-6.zip).
    """
    import re
    from urllib.parse import urljoin

    try:
        resp = requests.get(page_url, timeout=40, headers={"User-Agent": USER_AGENT})
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"✗ Could not load page {page_url}: {e}")
        return None

    hrefs = re.findall(r'href="([^"]+)"', resp.text)
    matches = [h for h in hrefs if name_contains in h and
               h.lower().endswith((".zip", ".xlsx", ".xls", ".csv"))]
    if not matches:
        print(f"✗ No '{name_contains}' file link found on {page_url}")
        return None

    url = urljoin(page_url, matches[0])
    print(f"✓ Found source file: {url}")
    return url


def _curl(args):
    """Run curl; some hosts (humanservices.hawaii.gov) break Python's TLS."""
    import subprocess
    return subprocess.run(["curl", "-sL", "-m", "120", "-A", USER_AGENT, *args],
                          capture_output=True)


def fetch_text(url):
    """GET a page as text, falling back to curl if Python's TLS can't handle the host."""
    try:
        r = requests.get(url, timeout=40, headers={"User-Agent": USER_AGENT})
        r.raise_for_status()
        return r.text
    except requests.exceptions.RequestException:
        res = _curl([url])
        return res.stdout.decode("utf-8", "replace") if res.returncode == 0 else None


def download_file(url, destination):
    """Download a file from URL to destination (curl fallback on TLS failure)."""
    print(f"Downloading from: {url}")
    print(f"Saving to: {destination}")
    destination.parent.mkdir(parents=True, exist_ok=True)

    try:
        response = requests.get(url, stream=True, timeout=120,
                                headers={"User-Agent": USER_AGENT})
        response.raise_for_status()
        with open(destination, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
    except requests.exceptions.RequestException as e:
        res = _curl(["-o", str(destination), url])
        if res.returncode != 0 or not destination.exists() or destination.stat().st_size == 0:
            print(f"✗ Download failed (requests: {e}; curl rc={res.returncode})")
            return False

    print(f"✓ Downloaded: {destination.name} ({destination.stat().st_size / 1024:.1f} KB)")
    return True


def update_monthly_data():
    """
    Update statewide monthly SNAP data.

    Current: FY89 - January 2022
    Target: FY89 - May 2025
    """
    print("\n" + "="*60)
    print("UPDATING STATEWIDE MONTHLY SNAP DATA")
    print("="*60)

    file_path = DATA_DIR / FILES["monthly"]

    # 1. Find and download the current per-FY ZIP straight from USDA.
    url = discover_file_url(URLS["monthly"], SOURCE_FILE_PATTERNS["monthly"])
    if not url:
        print("✗ Could not locate the monthly source file. Aborting.")
        return

    download_path = Path("downloads/snap-zip-fy69tocurrent.zip")
    if not download_file(url, download_path):
        return

    # 2. Extract Hawaii monthly records from every FY file in the ZIP.
    try:
        extracted = _extract_hawaii_monthly(download_path)
    except Exception as e:
        print(f"✗ Extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return

    if extracted is None or extracted.empty:
        print("✗ No Hawaii records extracted from the downloaded ZIP. Aborting.")
        return
    print(f"✓ Extracted {len(extracted)} Hawaii monthly records "
          f"({extracted['Date'].min().date()} to {extracted['Date'].max().date()})")

    # 3. Compare against existing data — only ever move FORWARD.
    existing_df = pd.read_csv(file_path)
    existing_df['Date'] = pd.to_datetime(existing_df['Date'])
    last_date = existing_df['Date'].max()
    print(f"  Existing data ends: {last_date.date()}")

    new_records = extracted[extracted['Date'] > last_date]
    if new_records.empty:
        print("✓ No new months available upstream — local data is already current.")
        print("  (USDA's published file is not newer than what you have; "
              "nothing written.)")
        return

    # 4. New data exists: back up, merge forward, save.
    backup_existing_data(file_path)
    updated_df = pd.concat([existing_df, new_records], ignore_index=True)
    updated_df = updated_df.drop_duplicates(subset=['Date'], keep='last')
    updated_df = updated_df.sort_values('Date')
    updated_df.to_csv(file_path, index=False)

    print(f"✓ Added {len(new_records)} new months "
          f"(through {new_records['Date'].max().date()})")
    print(f"  Saved: {file_path}  ({len(updated_df)} total records)")

    # 5. Stamp provenance so the source is self-documenting from here on.
    import re as _re
    rev_match = _re.search(r'-(\d+)\.zip$', url)
    write_provenance("monthly", {
        "dataset": FILES["monthly"],
        "publisher": "USDA/FNS",
        "sourcePage": URLS["monthly"],
        "sourceFile": url.rsplit("/", 1)[-1],
        "revision": f"-{rev_match.group(1)}" if rev_match else None,
        "dataAsOf": getattr(extracted, "attrs", {}).get("asOf"),
        "downloadedAt": datetime.now().strftime("%Y-%m-%d"),
        "extractedBy": "scripts/extract_hawaii_snap.py",
        "coverageStart": str(updated_df['Date'].min().date()),
        "coverageEnd": str(updated_df['Date'].max().date()),
        "provenanceSource": "automated-pull",
    })
    print("  → Re-run scripts/prepare_web_data.py to refresh the dashboard JSON.")


def _extract_hawaii_monthly(zip_path):
    """Extract a clean Hawaii monthly DataFrame from the per-FY ZIP.

    Reuses the parsing logic in extract_hawaii_snap.py so there is a single
    source of truth for the (quirky) USDA Excel layouts.
    """
    import glob
    import tempfile
    from extract_hawaii_snap import (extract_hawaii_from_fy_file,
                                      parse_month_to_date)

    with zipfile.ZipFile(zip_path) as zf:
        tmpdir = tempfile.mkdtemp(prefix="snap_fy_")
        zf.extractall(tmpdir)

    files = sorted(glob.glob(f"{tmpdir}/**/FY*.xls", recursive=True)) + \
        sorted(glob.glob(f"{tmpdir}/**/FY*.xlsx", recursive=True))

    records = []
    as_of = None
    for f in files:
        recs = extract_hawaii_from_fy_file(f)
        if recs:
            records.extend(recs)
            as_of = read_asof_date(f) or as_of  # latest FY file's "as of" date

    if not records:
        return None

    df = pd.DataFrame(records)
    df['Date'] = pd.to_datetime(df['Month'].apply(parse_month_to_date))
    df = df[df['Date'].notna()]
    df = df[['Date', 'Household', 'Persons', 'Per Household', 'Per Person', 'Cost']]
    df = df.sort_values('Date')
    df.attrs["asOf"] = as_of  # carried to the provenance stamp
    return df


def update_retailer_data():
    """
    Update historical SNAP retailer data.

    Current: 1990-2021
    Target: 1990 - December 31, 2024
    """
    print("\n" + "="*60)
    print("UPDATING SNAP RETAILER HISTORICAL DATA")
    print("="*60)

    # Find the most recent existing Hawaii retailer file (name carries a year
    # range, e.g. 2004-2024 → 2005-2025, so glob rather than hardcode).
    import glob as _glob
    prior = sorted(_glob.glob(str(DATA_DIR / "hawaii_snap_retailers_*_valid_coords.csv")))
    existing_path = Path(prior[-1]) if prior else DATA_DIR / FILES["retailers"]
    existing_count = 0
    if existing_path.exists():
        try:
            existing_count = len(pd.read_csv(existing_path))
        except Exception:
            pass
    print(f"Existing Hawaii retailer records: {existing_count} "
          f"({existing_path.name if existing_count else 'none'})")

    # 1. Find and download the current retailer ZIP straight from USDA.
    url = discover_file_url(URLS["retailers"], SOURCE_FILE_PATTERNS["retailers"])
    if not url:
        print("✗ Could not locate the retailer source file. Aborting.")
        return

    download_path = Path("downloads/snap-retailer-locator-data.zip")
    if not download_file(url, download_path):
        return

    # 2. Extract the CSV from the ZIP and filter to Hawaii.
    try:
        import tempfile, glob
        tmpdir = tempfile.mkdtemp(prefix="snap_retail_")
        with zipfile.ZipFile(download_path) as zf:
            zf.extractall(tmpdir)
        csvs = glob.glob(f"{tmpdir}/**/*.csv", recursive=True)
        if not csvs:
            print("✗ No CSV found inside the retailer ZIP. Aborting.")
            return
        # utf-8-sig strips the byte-order mark USDA prepends (else the first
        # column header comes through as "﻿Record ID").
        df = pd.read_csv(csvs[0], encoding="utf-8-sig", low_memory=False)
    except Exception as e:
        print(f"✗ Could not read retailer ZIP: {e}")
        return

    state_col = "State" if "State" in df.columns else (
        "store_state" if "store_state" in df.columns else None)
    if not state_col:
        print(f"✗ No State column. Columns: {df.columns.tolist()[:10]}")
        return
    hi_new = df[df[state_col] == "HI"].copy()
    print(f"✓ Filtered for Hawaii: {len(hi_new)} records (this release)")

    # Normalize date columns to ISO so keys/output are format-consistent. The
    # raw USDA file uses M/D/YYYY while prior repo files are already ISO — without
    # this, the merge key never matches and every store gets duplicated.
    def _norm_dates(frame):
        for col in ("Authorization Date", "End Date"):
            if col in frame.columns:
                frame[col] = pd.to_datetime(frame[col], errors="coerce").dt.strftime("%Y-%m-%d")
        return frame
    hi_new = _norm_dates(hi_new)

    # 3. UNION with existing — retailer data is a cumulative history, and each
    # USDA release spans only a rolling ~20-year window, so a plain replace would
    # silently drop stores that closed before the window (e.g. Hawaii stores that
    # closed in 2004). A store appears once per authorization period, so
    # (Record ID, Authorization Date) is the true primary key.
    key = [c for c in ["Record ID", "Authorization Date"] if c in hi_new.columns]
    prior_all = sorted(_glob.glob(str(DATA_DIR / "hawaii_snap_retailers_*_all.csv")))
    if prior_all and key:
        try:
            old_df = _norm_dates(pd.read_csv(prior_all[-1], encoding="utf-8-sig",
                                             low_memory=False))
            before = len(hi_new)
            # New release wins on conflicts (keep='last'); old-only rows survive.
            hi_df = pd.concat([old_df, hi_new], ignore_index=True)
            hi_df = hi_df.drop_duplicates(subset=key, keep="last")
            preserved = len(hi_df) - before
            print(f"  Unioned with {prior_all[-1].split('/')[-1]}: "
                  f"+{max(preserved,0)} historical record(s) the new release dropped "
                  f"(total {len(hi_df)})")
        except Exception as e:
            print(f"  ⚠ Could not union with prior file ({e}); using this release only")
            hi_df = hi_new
    else:
        hi_df = hi_new

    # Regression guard: the union must never be smaller than what we had.
    if existing_count and len(hi_df) < existing_count:
        print(f"✗ Union ({len(hi_df)}) smaller than existing ({existing_count}). "
              f"Refusing to regress; nothing written.")
        return

    # 4. Coordinate validation → write both an "all" file and a valid-coords file.
    lat_lo, lat_hi = HI_BOUNDS["lat"]
    lon_lo, lon_hi = HI_BOUNDS["lon"]
    if "Latitude" in hi_df.columns and "Longitude" in hi_df.columns:
        valid = ((hi_df["Latitude"] >= lat_lo) & (hi_df["Latitude"] <= lat_hi) &
                 (hi_df["Longitude"] >= lon_lo) & (hi_df["Longitude"] <= lon_hi))
        print(f"  Valid Hawaii coordinates: {valid.sum()}/{len(hi_df)} "
              f"({valid.sum()/len(hi_df)*100:.1f}%)")
    else:
        valid = pd.Series(True, index=hi_df.index)

    # Derive the coverage span from the actual data (End Date years) so the
    # filename reflects what's really in the unioned file, not the USDA window.
    ed_years = pd.to_datetime(hi_df.get("End Date"), errors="coerce").dt.year.dropna()
    if len(ed_years):
        span = f"{int(ed_years.min())}-{int(ed_years.max())}"
    else:
        span = "latest"
    all_path = DATA_DIR / f"hawaii_snap_retailers_{span}_all.csv"
    valid_path = DATA_DIR / f"hawaii_snap_retailers_{span}_valid_coords.csv"

    backup_existing_data(existing_path)
    hi_df.to_csv(all_path, index=False)
    hi_df[valid].to_csv(valid_path, index=False)
    print(f"✓ Saved {all_path.name} ({len(hi_df)}) and "
          f"{valid_path.name} ({valid.sum()})")

    # 5. Stamp provenance.
    write_provenance("retailers", {
        "dataset": valid_path.name,
        "publisher": "USDA/FNS",
        "sourcePage": URLS["retailers"],
        "sourceFile": url.rsplit("/", 1)[-1],
        "downloadedAt": datetime.now().strftime("%Y-%m-%d"),
        "coverageStart": span.split("-")[0],
        "coverageEnd": span.split("-")[-1],
        "hawaiiRecords": int(len(hi_df)),
        "validCoordRecords": int(valid.sum()),
        "note": ("Unioned across USDA releases to preserve history beyond USDA's "
                 "rolling ~20-year window (which drops stores that closed long ago)."),
        "provenanceSource": "automated-pull",
    })
    print("  → Re-run scripts/prepare_web_data.py to refresh the dashboard JSON.")


def update_county_data():
    """
    Update county bi-annual SNAP data.

    Current: FY89 - July 2020 (README claims Jan 2021 but file shows July 2020)
    Target: FY89 - latest available
    """
    print("\n" + "="*60)
    print("UPDATING COUNTY BI-ANNUAL SNAP DATA")
    print("="*60)

    file_path = DATA_DIR / FILES["county"]
    backup_existing_data(file_path)

    # Check current data
    try:
        existing_df = pd.read_csv(file_path)
        existing_df['Date'] = pd.to_datetime(existing_df['Date'])
        print(f"Current data: {existing_df['Date'].min()} to {existing_df['Date'].max()}")
        print(f"Records: {len(existing_df)}")
    except Exception as e:
        print(f"Error reading existing data: {e}")

    # County bi-annual (Jan/Jul) project-area data is no longer a direct file
    # link on the SNAP data page, so it can't be auto-discovered like the others.
    # Skip non-interactively (so --all never blocks) with a clear pointer.
    print("\n⚠ SKIPPED — no auto-discoverable source.")
    print("  The bi-annual county/project-area file isn't published as a direct")
    print("  download on the SNAP data page. To update manually, obtain the")
    print("  'Bi-Annual (Jan & Jul) State Project Area/County Level' data from")
    print("  USDA FNS, save the Hawaii rows in the existing CSV's schema, and")
    print("  append periods after "
          f"{existing_df['Date'].max().date() if 'existing_df' in dir() else 'the latest date'}.")


def generate_summary_report(output_path=None):
    """Generate a summary report of all datasets."""
    if output_path is None:
        output_path = DATA_DIR.parent / "DATA_STATUS_REPORT.md"

    print("\n" + "="*60)
    print("GENERATING DATA STATUS REPORT")
    print("="*60)

    report = ["# Hawaii SNAP Data Status Report"]
    report.append(f"\n**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    report.append("---\n")

    # Check each dataset
    for key, filename in FILES.items():
        report.append(f"\n## {key.replace('_', ' ').title()}")
        file_path = DATA_DIR / filename

        if not file_path.exists():
            report.append(f"\n⚠️ **Status:** File not found")
            report.append(f"\n- Expected path: `{file_path}`")
            continue

        try:
            df = pd.read_csv(file_path)
            report.append(f"\n✓ **Status:** Found")
            report.append(f"\n- **Records:** {len(df):,}")
            report.append(f"\n- **File size:** {file_path.stat().st_size / 1024:.1f} KB")
            report.append(f"\n- **Last modified:** {datetime.fromtimestamp(file_path.stat().st_mtime).strftime('%Y-%m-%d')}")

            # Date range
            date_columns = [col for col in df.columns if 'date' in col.lower()]
            if date_columns:
                date_col = date_columns[0]
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                report.append(f"\n- **Date range:** {df[date_col].min()} to {df[date_col].max()}")

            # Column info
            report.append(f"\n- **Columns:** {', '.join(df.columns.tolist())}")

        except Exception as e:
            report.append(f"\n⚠️ **Status:** Error reading file")
            report.append(f"\n- Error: {str(e)}")

    report_text = "\n".join(report)

    with open(output_path, 'w') as f:
        f.write(report_text)

    print(f"✓ Report saved: {output_path}")
    print("\n" + report_text)


def main():
    parser = argparse.ArgumentParser(
        description="Download and update Hawaii SNAP data",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('--monthly', action='store_true', help='Update statewide monthly data')
    parser.add_argument('--retailers', action='store_true', help='Update retailer historical data')
    parser.add_argument('--county', action='store_true', help='Update county bi-annual data')
    parser.add_argument('--dhs', action='store_true', help='Update Hawaii DHS participation + timeliness')
    parser.add_argument('--dhs-backfill', action='store_true', help='Rebuild DHS history from the archive (one-time)')
    parser.add_argument('--all', action='store_true', help='Update all datasets')
    parser.add_argument('--report', action='store_true', help='Generate status report only')

    args = parser.parse_args()

    if args.report:
        generate_summary_report()
        return

    if getattr(args, 'dhs_backfill', False):
        backfill_dhs_data()
        return

    if not any([args.monthly, args.retailers, args.county, args.dhs, args.all]):
        parser.print_help()
        print("\n" + "="*60)
        print("QUICK START:")
        print("="*60)
        print("1. Run with --report to see current data status")
        print("2. Run with --all to update all datasets")
        print("3. Follow manual download instructions for each dataset")
        print("\nExample: python download_and_update.py --all")
        return

    # Create necessary directories
    DATA_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    Path("downloads").mkdir(exist_ok=True)

    print("\n" + "="*60)
    print("HAWAII SNAP DATA UPDATE UTILITY")
    print("="*60)
    print(f"Data directory: {DATA_DIR.absolute()}")
    print(f"Backup directory: {BACKUP_DIR.absolute()}\n")

    # Update datasets
    if args.all or args.monthly:
        update_monthly_data()

    if args.all or args.retailers:
        update_retailer_data()

    if args.all or args.county:
        update_county_data()

    if args.all or args.dhs:
        update_dhs_data()

    # Rebuild the dashboard JSON so a full run updates data AND site together.
    if args.all or args.monthly:
        rebuild_web_data()

    # Generate final report
    print("\n" + "="*60)
    generate_summary_report()

    print("\n" + "="*60)
    print("UPDATE COMPLETE")
    print("="*60)
    print("\nNext steps:")
    print("1. Review DATA_STATUS_REPORT.md and Data/SOURCES.json (provenance)")
    print("2. git add/commit/push  → Netlify redeploys the dashboard automatically")


def _dhs_latest_link(page_text, fname_regex):
    """Return the (year, absolute_url) with the highest year matching a DHS file."""
    import re
    from urllib.parse import urljoin
    best = None
    for href in re.findall(r'href="([^"]+)"', page_text):
        m = re.search(fname_regex, href, re.IGNORECASE)
        if m:
            year = int(m.group(1))
            if best is None or year > best[0]:
                best = (year, urljoin(URLS["dhs"], href))
    return best


def update_dhs_data():
    """Pull Hawaii DHS monthly participation (by island) + application timeliness."""
    print("\n" + "="*60)
    print("UPDATING HAWAII DHS SNAP DATA (participation + timeliness)")
    print("="*60)

    page = fetch_text(URLS["dhs"])
    if not page:
        print("✗ Could not load DHS page.")
        return

    # Current machine-readable releases (PDF-era SFY/FFY ~2016-2024 not handled).
    part_link = _dhs_latest_link(page, r'Partic.*SFY[-\s]*(\d{4}).*\.xlsx?$')
    tl_link = _dhs_latest_link(page, r'Timeliness.*FFY[-\s]*(\d{4}).*\.xlsx?$')
    if not part_link or not tl_link:
        print(f"✗ Could not find current DHS files "
              f"(participation={bool(part_link)}, timeliness={bool(tl_link)}).")
        return

    part_path = Path("downloads/dhs_participation.xlsx")
    tl_path = Path("downloads/dhs_timeliness.xlsx")
    if not download_file(part_link[1], part_path) or not download_file(tl_link[1], tl_path):
        return

    from extract_dhs_snap import extract_participation, extract_timeliness
    part = pd.DataFrame(extract_participation(part_path))
    tl = pd.DataFrame(extract_timeliness(tl_path)).sort_values('Date')
    if part.empty or tl.empty:
        print("✗ DHS extraction produced no rows. Aborting.")
        return

    part_csv = DATA_DIR / "dhs_snap_participation_by_island.csv"
    tl_csv = DATA_DIR / "dhs_snap_application_timeliness.csv"
    # Union with existing history (the multi-year backfill) so monthly pulls
    # only ever add new months, never drop the archive. Keyed on Date+Island.
    if part_csv.exists():
        prior = pd.read_csv(part_csv)
        part = pd.concat([prior, part], ignore_index=True)
    part = part.drop_duplicates(['Date', 'Island'], keep='last').sort_values(['Date', 'Island'])
    if tl_csv.exists():
        tl = pd.concat([pd.read_csv(tl_csv), tl], ignore_index=True)
    tl = tl.drop_duplicates(['Date'], keep='last').sort_values('Date')
    part.to_csv(part_csv, index=False)
    tl.to_csv(tl_csv, index=False)
    print(f"✓ {part_csv.name}: {len(part)} rows ({part['Date'].min()}..{part['Date'].max()})")
    print(f"✓ {tl_csv.name}: {len(tl)} rows ({tl['Date'].min()}..{tl['Date'].max()})")

    # A routine pull only refreshes volatile fields; note/fiscalYear/coverageStart
    # are owned by backfill_dhs_data() and preserved (None values are skipped).
    write_provenance("dhs_participation", {
        "dataset": part_csv.name,
        "publisher": "Hawaii DHS (Department of Human Services)",
        "sourcePage": URLS["dhs"],
        "latestSourceFile": part_link[1].rsplit("/", 1)[-1],
        "downloadedAt": datetime.now().strftime("%Y-%m-%d"),
        "coverageEnd": str(part['Date'].max()),
        "granularity": "monthly, by island/branch",
    })
    write_provenance("dhs_timeliness", {
        "dataset": tl_csv.name,
        "publisher": "Hawaii DHS (Department of Human Services)",
        "sourcePage": URLS["dhs"],
        "latestSourceFile": tl_link[1].rsplit("/", 1)[-1],
        "downloadedAt": datetime.now().strftime("%Y-%m-%d"),
        "coverageEnd": str(tl['Date'].max()),
        "granularity": "monthly, statewide",
    })


def backfill_dhs_data():
    """One-time: rebuild DHS participation + timeliness history from the archives."""
    from extract_dhs_snap import (DHS_PARTICIPATION_ARCHIVE, DHS_TIMELINESS_ARCHIVE,
                                  extract_participation_any, extract_timeliness_any)

    print("\n" + "="*60)
    print("BACKFILLING DHS PARTICIPATION HISTORY (archive)")
    print("="*60)
    recs = []
    for sfy, url in sorted(DHS_PARTICIPATION_ARCHIVE.items()):
        dest = Path(f"downloads/dhs_archive/SFY-{sfy}.{url.rsplit('.', 1)[-1]}")
        if not download_file(url, dest):
            print(f"  ⚠ SFY {sfy}: download failed, skipping"); continue
        r = extract_participation_any(dest, sfy)
        print(f"  SFY {sfy}: {len(set(x['Date'] for x in r))} months")
        recs += r
    part_csv = DATA_DIR / "dhs_snap_participation_by_island.csv"
    df = pd.DataFrame(recs)
    if part_csv.exists():
        df = pd.concat([pd.read_csv(part_csv), df], ignore_index=True)
    df = df.drop_duplicates(['Date', 'Island'], keep='first').sort_values(['Date', 'Island'])
    df.to_csv(part_csv, index=False)
    st = df[df['Island'] == 'STATE']
    print(f"✓ {part_csv.name}: {len(df)} rows, STATE {st['Date'].min()}..{st['Date'].max()} ({len(st)} months)")

    print("\n" + "="*60)
    print("BACKFILLING DHS TIMELINESS HISTORY (archive)")
    print("="*60)
    trecs = []
    for ffy, url in sorted(DHS_TIMELINESS_ARCHIVE.items()):
        dest = Path(f"downloads/dhs_archive/TL-{ffy}.{url.rsplit('.', 1)[-1]}")
        if not download_file(url, dest):
            print(f"  ⚠ FFY {ffy}: download failed, skipping"); continue
        r = extract_timeliness_any(dest, ffy) or []
        print(f"  FFY {ffy}: {len(r)} months")
        trecs += r
    tl_csv = DATA_DIR / "dhs_snap_application_timeliness.csv"
    tdf = pd.DataFrame(trecs)
    if tl_csv.exists():
        tdf = pd.concat([pd.read_csv(tl_csv), tdf], ignore_index=True)
    tdf = tdf.drop_duplicates(['Date'], keep='first').sort_values('Date')
    tdf.to_csv(tl_csv, index=False)
    print(f"✓ {tl_csv.name}: {len(tdf)} months, {tdf['Date'].min()}..{tdf['Date'].max()}")

    # Record the descriptive provenance the backfill owns (routine --dhs pulls
    # preserve these; they only refresh downloadedAt/coverageEnd).
    write_provenance("dhs_participation", {
        "coverageStart": str(st['Date'].min()), "coverageEnd": str(st['Date'].max()),
        "fiscalYear": "SFY 2009-2026", "provenanceSource": "automated-pull + backfill",
        "note": "Monthly by-island participation, SFY 2009-2026. History backfilled "
                "from the DHS archive via scripts/extract_dhs_snap.py (.xls via xlrd, "
                "PDF via pdftotext). Cross-validated: the July-2021 peak (206,226) "
                "matches USDA's all-time max exactly, and island parts sum to STATE. "
                "May/June 2021 missing (the SFY2021 release was partial).",
    })
    write_provenance("dhs_timeliness", {
        "coverageStart": str(tdf['Date'].min()), "coverageEnd": str(tdf['Date'].max()),
        "fiscalYear": "FFY 2009-2026", "provenanceSource": "automated-pull + backfill",
        "note": "Application processing timeliness + applications received, FFY "
                "2009-2026. Backfilled from the DHS archive (.xls 'STATE SUMMARY' "
                "sheets via xlrd; PDF 'State TOTAL' rows via pdftotext). Remaining "
                "gaps are source-side: FFY2025 unpublished; FFY2018/2020/2021 are "
                "mid-year partials; FFY2023 is an August-only snapshot.",
    })


def rebuild_web_data():
    """Regenerate web/data/*.json so the dashboard reflects updated CSVs."""
    print("\n" + "="*60)
    print("REBUILDING DASHBOARD DATA (web/data/*.json)")
    print("="*60)
    import subprocess
    script = Path(__file__).parent / "prepare_web_data.py"
    result = subprocess.run([sys.executable, str(script)],
                            capture_output=True, text=True)
    if result.returncode == 0:
        print("✓ Dashboard JSON rebuilt")
    else:
        print("✗ prepare_web_data.py failed:")
        print(result.stdout[-1500:])
        print(result.stderr[-1500:])


if __name__ == "__main__":
    main()
