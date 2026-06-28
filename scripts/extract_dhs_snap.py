#!/usr/bin/env python3
"""
Extract Hawaii DHS SNAP data (monthly, by island/branch) into tidy CSVs.

Two ongoing DHS series, published on https://humanservices.hawaii.gov/bessd/snap/:
  - Participation Summary (SFY) -> participants/households/benefits by island
  - Application Timeliness (FFY) -> applications received + on-time disposition rates

These are richer than the USDA monthly series (island-level, program breakdown)
and more current. This module parses the machine-readable .xls/.xlsx releases.
PDF-era releases (roughly SFY/FFY 2016-2024) are not handled here.

Usage:
    python scripts/extract_dhs_snap.py PARTICIPATION.xlsx TIMELINESS.xlsx
"""

import re
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Island/branch codes (old .xls/PDF) and full names (2025+ .xlsx) -> canonical.
CANON_ISLAND = {
    'OB': 'Oahu Branch', 'OAHU BRANCH': 'Oahu Branch',
    'HB': 'Hawaii Branch', 'HAWAII BRANCH': 'Hawaii Branch',
    'KB': 'Kauai Branch', 'KAUAI BRANCH': 'Kauai Branch',
    'MAUI': 'Maui Island', 'MAUI ISLAND': 'Maui Island',
    'MOLOKAI': 'Molokai Island', 'MOLOKAI ISLAND': 'Molokai Island',
    'LANAI': 'Lanai Island', 'LANAI ISLAND': 'Lanai Island',
    'MB': 'Maui Branch', 'MAUI BRANCH': 'Maui Branch',
    'STATE': 'STATE', 'STATE TOTAL': 'STATE',
}
# Section titles for the three metrics, across eras.
METRIC_TITLES = {
    'Participants': ('NUMBER OF PERSONS', 'NUMBER OF PARTICIPANTS RECEIVING'),
    'Households': ('NUMBER OF HOUSEHOLDS', 'NUMBER OF HOUSEHOLDS RECEIVING'),
    'BenefitsIssued': ('COUPON ISSUANCE', 'SNAP BENEFITS ISSUED'),
}


def _sfy_year(month_idx, sfy):
    """State Fiscal Year N runs Jul(N-1)..Jun(N)."""
    return sfy - 1 if month_idx >= 7 else sfy


def _fy_from_name(path, kind):
    """Pull SFY/FFY year from a filename like 'SFY-2018-...' or 'FFY-2020'."""
    m = re.search(rf'{kind}[-_\s]*(\d{{4}})', Path(path).name, re.IGNORECASE)
    return int(m.group(1)) if m else None

MONTHS = {m: i for i, m in enumerate(
    ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
     'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], start=1)}
MONTH_NAMES = {'JANUARY': 1, 'FEBRUARY': 2, 'MARCH': 3, 'APRIL': 4, 'MAY': 5,
               'JUNE': 6, 'JULY': 7, 'AUGUST': 8, 'SEPTEMBER': 9,
               'OCTOBER': 10, 'NOVEMBER': 11, 'DECEMBER': 12}
ISLANDS = {'Oahu Branch', 'Hawaii Branch', 'Kauai Branch', 'Maui Island',
           'Molokai Island', 'Lanai Island', 'Maui Branch', 'STATE'}
# Section 1 table titles -> output metric name
PART_TABLES = {
    'Number of Participants Receiving SNAP Benefits': 'Participants',
    'Number of Households Receiving SNAP Benefits': 'Households',
    'SNAP Benefits Issued': 'BenefitsIssued',
}


def _sheet_date(sheet_name):
    """'JUL 2025' -> '2025-07-01' (else None)."""
    m = re.match(r'([A-Za-z]{3})\s+(\d{4})', sheet_name.strip())
    if m and m.group(1).upper() in MONTHS:
        return f"{m.group(2)}-{MONTHS[m.group(1).upper()]:02d}-01"
    return None


def extract_participation(path):
    """Tidy monthly by-island participation records from a Participation .xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        date = _sheet_date(sheet)
        if not date:
            continue  # Title Page / Summary
        rows = list(wb[sheet].iter_rows(values_only=True))
        # collect each island's value per metric
        by_island = {}
        metric = None
        for r in rows:
            c0 = (str(r[0]).strip() if r and r[0] is not None else '')
            # Sections 2/3 repeat the island rows with a different 7-column
            # layout (no TOTAL); stop so they don't clobber Section 1 values.
            if c0.startswith('Section 2'):
                break
            matched = next((v for k, v in PART_TABLES.items() if c0.startswith(k)), None)
            if matched:
                metric = matched
                continue
            if metric and c0 in ISLANDS:
                total = r[7] if len(r) > 7 else None        # TOTAL column
                snap_only = r[6] if len(r) > 6 else None     # SNAP ONLY column
                rec = by_island.setdefault(c0, {})
                rec[metric] = total
                if metric == 'Participants':
                    rec['ParticipantsSNAPOnly'] = snap_only
        # a fully-zero STATE row means the month isn't reported yet
        state = by_island.get('STATE', {})
        if not state or not state.get('Participants'):
            continue
        for island, vals in by_island.items():
            out.append({
                'Date': date, 'Island': island,
                'Participants': vals.get('Participants'),
                'Households': vals.get('Households'),
                'BenefitsIssued': vals.get('BenefitsIssued'),
                'ParticipantsSNAPOnly': vals.get('ParticipantsSNAPOnly'),
            })
    return out


def _metric_for_title(line_upper):
    """Map a section-title line to a metric, or None. Guards section-2 lookalikes."""
    if 'CATEGORY' in line_upper or 'BY FA' in line_upper or 'BY AREA' in line_upper:
        return None
    for metric, titles in METRIC_TITLES.items():
        if any(t in line_upper for t in titles):
            return metric
    return None


def _island_row_values(tokens):
    """Given whitespace tokens of a data row, return (island, snap_only, total) or None.

    Columns are TANF TAONF GA SSI ABD NPA/SNAP-ONLY TOTAL — take the first 7
    integers; TOTAL is the 7th, NPA/SNAP-only the 6th. Trailing acronym-key text
    is ignored because we stop after 7 numbers.
    """
    is_num = lambda t: bool(re.match(r'^[\d,]+\.?\d*$', t)) and any(c.isdigit() for c in t)
    # island code = leading non-numeric, non-empty tokens (e.g. 'OB', 'Maui', 'STATE')
    lead = []
    i = 0
    while i < len(tokens) and not is_num(tokens[i]):
        if tokens[i].strip():
            lead.append(tokens[i].upper())
        i += 1
    key = ' '.join(lead).strip().replace(' TOTALS', '').replace(' TOTAL', '')
    island = CANON_ISLAND.get(key) or CANON_ISLAND.get(key.split()[0] if key else '', None)
    if not island:
        return None
    # Collect the full numeric run (column count varies by era: some years add
    # an 'NA' column). TOTAL is always the last column, NPA/SNAP-only the one
    # before it. Stop at the first trailing text token (acronym-key sidebar).
    nums = []
    for t in tokens[i:]:
        if is_num(t):
            nums.append(int(round(float(t.replace(',', '')))))
        elif t.strip():
            break
    if len(nums) < 6:
        return None
    return island, nums[-2], nums[-1]


def _participation_from_rows(get_rows, date_for_sheet):
    """Shared engine: iterate (sheet, rows-of-token-lists); pull the 3 metric tables."""
    out = {}
    for sheet, rows in get_rows():
        date = date_for_sheet(sheet)
        if not date:
            continue
        metric = None
        for tokens in rows:
            if not tokens:
                continue
            line_up = ' '.join(str(t) for t in tokens).upper()
            m = _metric_for_title(line_up)
            if m:
                metric = m
                continue
            if 'SECTION 2' in line_up or 'BY CATEGORY' in line_up:
                metric = None
            if not metric:
                continue
            vals = _island_row_values([str(t) for t in tokens])
            if vals:
                island, snap_only, total = vals
                rec = out.setdefault((date, island), {})
                rec[metric] = total
                if metric == 'Participants':
                    rec['ParticipantsSNAPOnly'] = snap_only
    return [
        {'Date': d, 'Island': isl, 'Participants': v.get('Participants'),
         'Households': v.get('Households'), 'BenefitsIssued': v.get('BenefitsIssued'),
         'ParticipantsSNAPOnly': v.get('ParticipantsSNAPOnly')}
        for (d, isl), v in out.items() if v.get('Participants')
    ]


def _participation_from_xls(path, sfy):
    """Old .xls files: month sheets JUL..JUN (bare), tables NUMBER OF PERSONS etc."""
    import xlrd
    wb = xlrd.open_workbook(path)

    def get_rows():
        for name in wb.sheet_names():
            mi = MONTHS.get(name.strip().upper()[:3])
            if not mi:
                continue
            sh = wb.sheet_by_name(name)
            rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
            yield (name, mi), rows

    return _participation_from_rows(
        get_rows, lambda key: f"{_sfy_year(key[1], sfy)}-{key[1]:02d}-01")


def _participation_from_pdf(path, sfy):
    """PDF summaries (~2016-2024): per-page month worksheets, same table layout."""
    txt = subprocess.run(["pdftotext", "-layout", str(path), "-"],
                         capture_output=True, text=True).stdout

    def get_rows():
        for page in txt.split('\f'):
            mi = None
            for line in page.splitlines()[:8]:
                pm = re.search(r'\((JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\)',
                               line.upper())
                if pm:
                    mi = MONTHS[pm.group(1)]
                    break
            if not mi:
                continue
            rows = [ln.split() for ln in page.splitlines()]
            yield (page, mi), rows

    return _participation_from_rows(
        get_rows, lambda key: f"{_sfy_year(key[1], sfy)}-{key[1]:02d}-01")


def _timeliness_from_pdf(path, ffy):
    """PDF timeliness (~2016-2024): per-month pages, sub-office rows + 'State TOTAL'.

    The 'State TOTAL' row gives, in order: apps received, all dispositions,
    timely dispositions, % timely (followed by untimely/expedited columns)."""
    txt = subprocess.run(["pdftotext", "-layout", str(path), "-"],
                         capture_output=True, text=True).stdout
    out = {}
    for page in txt.split('\f'):
        # month marker near the top — standalone on its own line in most files,
        # but embedded in the column header in some (e.g. FFY2018).
        mi = None
        for line in page.splitlines()[:6]:
            mm = re.search(r'\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b',
                           line.strip().upper())
            if mm:
                mi = MONTHS[mm.group(1)]
                break
        if not mi:
            continue
        for line in page.splitlines():
            if re.match(r'\s*STATE\s+TOTAL', line, re.IGNORECASE):
                ints = [int(x.replace(',', '')) for x in re.findall(r'[\d,]+(?=\s|$)', line)
                        if re.fullmatch(r'[\d,]+', x)]
                pcts = re.findall(r'(\d+(?:\.\d+)?)%', line)
                if len(ints) >= 3 and pcts:
                    year = ffy - 1 if mi >= 10 else ffy
                    out[f"{year}-{mi:02d}-01"] = {
                        'ApplicationsReceived': ints[0],
                        'TotalDispositions': ints[1],
                        'TimelyDispositions': ints[2],
                        'PercentTimely': round(float(pcts[0]) / 100, 3),
                    }
                break
    return [{'Date': d, **v} for d, v in out.items()]


def _timeliness_from_summary(path, ffy):
    """Old .xls/.xlsx timeliness: a 'STATE SUMMARY' sheet, month x metric table."""
    if str(path).lower().endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(path)
        if 'STATE SUMMARY' not in wb.sheet_names():
            return None
        sh = wb.sheet_by_name('STATE SUMMARY')
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'STATE SUMMARY' not in wb.sheetnames:
            return None
        rows = list(wb['STATE SUMMARY'].iter_rows(values_only=True))

    out = []
    for r in rows:
        mon = MONTH_NAMES.get(str(r[0]).strip().upper()) if r and r[0] else None
        if not mon:
            continue
        nums = [x for x in r[1:] if isinstance(x, (int, float))]
        if len(nums) < 4:
            continue
        year = ffy - 1 if mon >= 10 else ffy
        out.append({
            'Date': f"{year}-{mon:02d}-01",
            'ApplicationsReceived': int(nums[0]),
            'TotalDispositions': int(nums[1]),
            'TimelyDispositions': int(nums[2]),
            'PercentTimely': round(float(nums[3]), 3),   # already a fraction
        })
    return out


def extract_timeliness_any(path, ffy=None):
    """Dispatch a timeliness file to the right parser by extension/era."""
    p = str(path)
    ffy = ffy or _fy_from_name(path, 'FFY')
    if p.lower().endswith('.pdf'):
        return _timeliness_from_pdf(path, ffy)
    summary = _timeliness_from_summary(path, ffy)   # old 'STATE SUMMARY' layout
    if summary:
        return summary
    return extract_timeliness(path)   # current 'FFY YYYY'/Table-1 layout (2026+)


def extract_participation_any(path, sfy=None):
    """Dispatch a participation file to the right parser by extension/era."""
    p = str(path)
    if p.lower().endswith('.xlsx'):
        return extract_participation(path)          # 2025+ dated-sheet format
    sfy = sfy or _fy_from_name(path, 'SFY')
    if p.lower().endswith('.xls'):
        return _participation_from_xls(path, sfy)
    if p.lower().endswith('.pdf'):
        return _participation_from_pdf(path, sfy)
    return []


def extract_timeliness(path):
    """Tidy monthly statewide application-timeliness records from a Timeliness file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        m = re.search(r'FFY\s*(\d{4})', sheet)
        if not m:
            continue
        ffy = int(m.group(1))
        in_table1 = False
        for r in wb[sheet].iter_rows(values_only=True):
            c0 = (str(r[0]).strip() if r and r[0] is not None else '')
            if c0.startswith('Table 1'):
                in_table1 = True
                continue
            if c0.startswith('Table 2'):
                in_table1 = False
                continue
            if not in_table1:
                continue
            mon = MONTH_NAMES.get(c0.upper())
            if not mon:
                continue
            received = r[1] if len(r) > 1 else None
            if not received:  # unreported future month
                continue
            # FFY N runs Oct (N-1) .. Sep (N)
            year = ffy - 1 if mon >= 10 else ffy
            out.append({
                'Date': f"{year}-{mon:02d}-01",
                'ApplicationsReceived': received,
                'TotalDispositions': r[2] if len(r) > 2 else None,
                'TimelyDispositions': r[3] if len(r) > 3 else None,
                'PercentTimely': r[4] if len(r) > 4 else None,
            })
    return out


# Hawaii DHS archive of machine-readable participation files (SFY -> URL).
# Used for the one-time historical backfill; the live pipeline only needs the
# current release. PDF-era files (2016-2024) are parsed via pdftotext.
DHS_PARTICIPATION_ARCHIVE = {
    2009: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-rpt-deleted-SFY-2009-SUMMARY.xls",
    2010: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-rpt-deleted-SFY-2010-SUMMARY.xls",
    2011: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-rpt-deleted-SFY-2011-SUMMARY-3.xls",
    2012: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-Rpt-deleted-SFY-2012-SUMMARY.xls",
    2013: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-rpt-deleted-SFY-2013-SUMMARY.xls",
    2014: "https://humanservices.hawaii.gov/bessd/files/2014/09/Corrected-and-fed-rpt-deleted-SNAP_SFY-2014-SUMMARY.xls",
    2015: "https://humanservices.hawaii.gov/wp-content/uploads/2016/06/SFY-2015-SUMMARY-.xls",
    2016: "https://humanservices.hawaii.gov/wp-content/uploads/2016/12/SFY-2016-Participation-SUMMARY.pdf",
    2017: "https://humanservices.hawaii.gov/wp-content/uploads/2017/08/SFY-2017-SUMMARY-thru-June.pdf",
    2018: "https://humanservices.hawaii.gov/wp-content/uploads/2018/08/SNAP-Participation.SFY-2018-SUMMARY.-as-of-June.pdf",
    2019: "https://humanservices.hawaii.gov/wp-content/uploads/2020/06/SFY-2019-SUMMARY.pdf",
    2020: "https://humanservices.hawaii.gov/wp-content/uploads/2020/07/SFY-2020-PART-SUMMARY.pdf",
    2021: "https://humanservices.hawaii.gov/wp-content/uploads/2021/05/SFY-2021-SUMMARY.pdf",
    2022: "https://humanservices.hawaii.gov/wp-content/uploads/2022/12/SFY-2022-SUMMARY.pdf",
    2023: "https://humanservices.hawaii.gov/wp-content/uploads/2023/10/SFY-2023-SUMMARY.pdf",
    2024: "https://humanservices.hawaii.gov/wp-content/uploads/2025/03/SFY-2024-SUMMARY.pdf",
    2025: "https://humanservices.hawaii.gov/wp-content/uploads/2025/07/SFY-2025-SUMMARY.xls",
}


DHS_TIMELINESS_ARCHIVE = {
    2009: "https://humanservices.hawaii.gov/bessd/files/2015/04/Application-Timeliness-FFY-20091.xls",
    2010: "https://humanservices.hawaii.gov/bessd/files/2015/04/Application-Timeliness-FFY-2010.xls",
    2011: "https://humanservices.hawaii.gov/bessd/files/2015/04/Application-Timeliness-FFY-2011.xls",
    2012: "https://humanservices.hawaii.gov/bessd/files/2015/04/Application-Timeliness-FFY20121.xls",
    2013: "https://humanservices.hawaii.gov/bessd/files/2012/12/application-timeliness-FFY-2013-corrected1.xls",
    2014: "https://humanservices.hawaii.gov/wp-content/uploads/2016/06/Application-Timeliness-FFY-20141.xls",
    2015: "https://humanservices.hawaii.gov/bessd/files/2016/05/Application-Timeliness-FFY-2015-new1.xlsx",
    2016: "https://humanservices.hawaii.gov/wp-content/uploads/2016/12/Application-Timeliness-FFY-2016-Complete.pdf",
    2017: "https://humanservices.hawaii.gov/wp-content/uploads/2017/12/Application-Timeliness-FFY-2017.pdf",
    2018: "https://humanservices.hawaii.gov/wp-content/uploads/2018/08/SNAP-Timeliness.FFY-2018-SUMMARY.-as-of-June.pdf",
    2019: "https://humanservices.hawaii.gov/wp-content/uploads/2020/06/Application-Timeliness-complete-FFY-2019.pdf",
    2020: "https://humanservices.hawaii.gov/wp-content/uploads/2020/08/Application-Timeliness-FFY-2020.pdf",
    2021: "https://humanservices.hawaii.gov/bessd/files/2021/06/Application-Timeliness-FFY-2021.pdf",
    2022: "https://humanservices.hawaii.gov/wp-content/uploads/2022/12/Application-Timeliness-FFY-2022.pdf",
    2023: "https://humanservices.hawaii.gov/wp-content/uploads/2023/10/Application-Timeliness-FFY-2023-aug.pdf",
    2024: "https://humanservices.hawaii.gov/wp-content/uploads/2025/03/Application-Timeliness-FFY-2024-1.pdf",
    # FFY2021/2018/2020 are partial (published mid-year); FFY2023 source is
    # August-only; FFY2025 was never published.
}


def main():
    if len(sys.argv) < 3:
        print("Usage: extract_dhs_snap.py PARTICIPATION.xlsx TIMELINESS.xlsx")
        sys.exit(1)
    data_dir = Path(__file__).parent.parent / "Data"

    part = pd.DataFrame(extract_participation(sys.argv[1])).sort_values(['Date', 'Island'])
    part_path = data_dir / "dhs_snap_participation_by_island.csv"
    part.to_csv(part_path, index=False)
    print(f"✓ {part_path.name}: {len(part)} rows, "
          f"{part['Date'].min()}..{part['Date'].max()}")

    tl = pd.DataFrame(extract_timeliness(sys.argv[2])).sort_values('Date')
    tl_path = data_dir / "dhs_snap_application_timeliness.csv"
    tl.to_csv(tl_path, index=False)
    print(f"✓ {tl_path.name}: {len(tl)} rows, {tl['Date'].min()}..{tl['Date'].max()}")


if __name__ == "__main__":
    main()
